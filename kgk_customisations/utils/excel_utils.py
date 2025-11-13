# Copyright (c) 2025, Apjakal IT Solutions and contributors
# For license information, please see license.txt

"""
Excel Generation Utilities
==========================

Centralized functions for Excel workbook creation, formatting, and styling.
Used across multiple reports and data export features.
"""

import frappe
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import io
from frappe.utils import now_datetime, today


def create_styled_excel_workbook(sheets_data, filename_prefix="report"):
	"""
	Create an Excel workbook with multiple sheets and consistent styling.
	
	Args:
		sheets_data: Dict or list of sheet configurations
			Format: {
				"sheet_name": {
					"headers": [...],
					"data": [[...], [...]],
					"special_formatting": {...}
				}
			}
		filename_prefix: Base name for the file
		
	Returns:
		Dict with file info and download details
	"""
	try:
		from openpyxl import Workbook
		from openpyxl.styles import PatternFill, Font
		
		# Create workbook
		wb = Workbook()
		
		# Remove default sheet
		wb.remove(wb.active)
		
		# Create sheets
		for sheet_name, sheet_config in sheets_data.items():
			ws = wb.create_sheet(title=sheet_name)
			
			# Add headers
			if sheet_config.get("headers"):
				ws.append(sheet_config["headers"])
				apply_standard_header_formatting(ws, 1, sheet_config.get("special_formatting"))
			
			# Add data
			if sheet_config.get("data"):
				for row in sheet_config["data"]:
					ws.append(row)
			
			# Apply special formatting if specified
			if sheet_config.get("special_formatting"):
				apply_special_formatting(ws, sheet_config["special_formatting"])
		
		# Generate filename
		filename = f"{filename_prefix}_{today().replace('-', '_')}.xlsx"
		
		return {
			"workbook": wb,
			"filename": filename,
			"success": True
		}
		
	except Exception as e:
		frappe.log_error(f"Error creating Excel workbook: {str(e)}")
		return {
			"workbook": None,
			"filename": None,
			"success": False,
			"error": str(e)
		}


def apply_standard_header_formatting(worksheet, header_row=1, special_columns=None):
	"""
	Apply consistent header formatting across all Excel reports.
	
	Args:
		worksheet: Openpyxl worksheet object
		header_row: Row number containing headers (default 1)
		special_columns: Dict of column ranges with special formatting
		
	Returns:
		Formatted worksheet
	"""
	try:
		# Standard header formatting
		header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
		header_font = Font(bold=True)
		
		# Apply to all header cells
		for cell in worksheet[header_row]:
			cell.fill = header_fill
			cell.font = header_font
		
		# Apply special formatting if specified
		if special_columns:
			for column_type, column_info in special_columns.items():
				if column_type == "highlighted_columns":
					# Apply special highlighting to specific columns
					highlight_fill = PatternFill(
						start_color=column_info.get("color", "FFEB9C"), 
						end_color=column_info.get("color", "FFEB9C"), 
						fill_type="solid"
					)
					
					for col_idx in column_info.get("columns", []):
						cell = worksheet.cell(row=header_row, column=col_idx)
						cell.fill = highlight_fill
						cell.font = header_font
		
		return worksheet
		
	except Exception as e:
		frappe.log_error(f"Error applying header formatting: {str(e)}")
		return worksheet


def apply_special_formatting(worksheet, formatting_config):
	"""
	Apply special formatting to worksheet based on configuration.
	
	Args:
		worksheet: Openpyxl worksheet object
		formatting_config: Dict with formatting instructions
	"""
	try:
		if "highlighted_columns" in formatting_config:
			config = formatting_config["highlighted_columns"]
			highlight_fill = PatternFill(
				start_color=config.get("color", "FFF2CC"), 
				end_color=config.get("color", "FFF2CC"), 
				fill_type="solid"
			)
			
			# Apply to data rows (skip header)
			start_row = config.get("start_row", 2)
			end_row = worksheet.max_row
			
			for col_idx in config.get("columns", []):
				for row_idx in range(start_row, end_row + 1):
					cell = worksheet.cell(row=row_idx, column=col_idx)
					cell.fill = highlight_fill
		
		# Set column widths if specified
		if "column_widths" in formatting_config:
			for col_letter, width in formatting_config["column_widths"].items():
				worksheet.column_dimensions[col_letter].width = width
		
	except Exception as e:
		frappe.log_error(f"Error applying special formatting: {str(e)}")


def create_summary_sheet(workbook, summary_data, sheet_name="Summary"):
	"""
	Create a summary sheet with key metrics and statistics.
	
	Args:
		workbook: Openpyxl workbook object
		summary_data: Dict with summary information
		sheet_name: Name for the summary sheet
		
	Returns:
		Summary worksheet
	"""
	try:
		# Create summary sheet
		summary_ws = workbook.create_sheet(title=sheet_name, index=0)  # Insert at beginning
		
		# Add title
		summary_ws['A1'] = "Report Summary"
		summary_ws['A1'].font = Font(bold=True, size=16)
		
		row = 3  # Start from row 3
		
		# Add summary data
		for key, value in summary_data.items():
			summary_ws[f'A{row}'] = key
			summary_ws[f'B{row}'] = value
			summary_ws[f'A{row}'].font = Font(bold=True)
			row += 1
		
		# Set column widths
		summary_ws.column_dimensions['A'].width = 25
		summary_ws.column_dimensions['B'].width = 15
		
		return summary_ws
		
	except Exception as e:
		frappe.log_error(f"Error creating summary sheet: {str(e)}")
		return None


def save_excel_as_frappe_file(workbook, filename, is_private=True, folder="Home"):
	"""
	Save Excel workbook as Frappe File document for download.
	
	Args:
		workbook: Openpyxl workbook object
		filename: Desired filename
		is_private: Whether file should be private
		folder: Frappe folder location
		
	Returns:
		File document with download URL
	"""
	try:
		# Save workbook to bytes
		excel_buffer = io.BytesIO()
		workbook.save(excel_buffer)
		excel_buffer.seek(0)
		
		# Create Frappe File document
		file_doc = frappe.get_doc({
			"doctype": "File",
			"file_name": filename,
			"is_private": is_private,
			"folder": folder,
			"content": excel_buffer.getvalue()
		})
		file_doc.insert()
		frappe.db.commit()
		
		return {
			"success": True,
			"file_doc": file_doc,
			"file_url": file_doc.file_url,
			"filename": filename
		}
		
	except Exception as e:
		frappe.log_error(f"Error saving Excel file: {str(e)}")
		return {
			"success": False,
			"error": str(e),
			"file_doc": None,
			"file_url": None
		}


def create_download_response(workbook, filename):
	"""
	Create direct download response for Excel file.
	
	Args:
		workbook: Openpyxl workbook object  
		filename: Desired filename
		
	Returns:
		Sets up Frappe response for direct download
	"""
	try:
		# Save workbook to buffer
		excel_buffer = io.BytesIO()
		workbook.save(excel_buffer)
		excel_buffer.seek(0)
		
		# Set up download response
		frappe.local.response.filename = filename
		frappe.local.response.filecontent = excel_buffer.getvalue()
		frappe.local.response.type = "download"
		frappe.local.response.headers = {
			"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			"Content-Disposition": f"attachment; filename={filename}"
		}
		
		return {"success": True, "message": "Download initiated"}
		
	except Exception as e:
		frappe.log_error(f"Error creating download response: {str(e)}")
		return {"success": False, "error": str(e)}