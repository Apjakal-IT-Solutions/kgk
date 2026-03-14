# Copyright (c) 2026, Apjakal IT Solutions and contributors
# For license information, please see license.txt

import io
import os
import shutil
import frappe
from frappe.model.document import Document
from frappe.utils import now_datetime, getdate, cint

_MOUNT_BASE = "/mnt/share/e-dox/Documents"


def _url_to_path(url):
	"""Resolve any file URL stored in a Cash Document field to an absolute path.

	Handles three origins:
	  /edox/...          → network mount (legacy migrated documents)
	  /private/files/... → Frappe private upload (new documents)
	  /files/...         → Frappe public upload (new documents)

	Returns None if the URL does not match a known scheme.
	"""
	if not url:
		return None
	if url.startswith("/edox/"):
		return os.path.join(_MOUNT_BASE, url[len("/edox/"):])
	if url.startswith("/private/files/"):
		return frappe.get_site_path("private", "files", url[len("/private/files/"):])
	if url.startswith("/files/"):
		return frappe.get_site_path("public", "files", url[len("/files/"):])
	return None

# Counter field names in the Cash Voucher Series Single DocType
_COUNTER_MAP = {
	"Cash":   "cash_counter",
	"Bank":   "bank_counter",
	"Cash-2": "cash_2_counter",
	"Bank-2": "bank_2_counter",
	"JE":     "je_counter",
}

# Name format strings (one positional placeholder for the counter)
_FORMAT_MAP = {
	"Cash":   "25-{}",
	"Bank":   "B{}",
	"Cash-2": "C2-25-{}",
	"Bank-2": "B2-{}",
	"JE":     "JE{}",
}

# Allowed sub_type values per main_type (empty string means no sub_type is OK)
_VALID_SUB_TYPES = {
	"Cash":   {"Payment", "Receipt"},
	"Bank":   {"Credit Card", "EFT"},
	"Cash-2": {""},
	"Bank-2": {""},
	"JE":     {"JE"},
}

class CashDocument(Document):

	def autoname(self):
		"""Assign a name derived from an atomic per-type counter."""
		main_type = self.main_type
		if main_type not in _COUNTER_MAP:
			frappe.throw(frappe._("Select a Document Type before saving."))

		field = _COUNTER_MAP[main_type]

		# Atomic increment using UPDATE.  tabSingles has no UNIQUE constraint, so
		# INSERT ... ON DUPLICATE KEY UPDATE creates duplicate rows and cannot be used.
		# The counter row is guaranteed to exist (seeded at install time by the patch).
		# InnoDB row-level locking on the UPDATE makes this safe under concurrent access.
		frappe.db.sql(
			"""
			UPDATE `tabSingles`
			SET value = CAST(COALESCE(CAST(value AS UNSIGNED), 0) + 1 AS CHAR)
			WHERE doctype = 'Cash Voucher Series' AND field = %(field)s
			""",
			{"field": field},
		)

		# Read back via raw SQL to bypass Frappe's Single-doctype in-memory cache.
		# InnoDB read-your-own-writes guarantee means we see the updated value
		# within the same transaction.
		result = frappe.db.sql(
			"SELECT value FROM `tabSingles` WHERE doctype='Cash Voucher Series' AND field=%s",
			[field],
		)
		next_num = cint(result[0][0]) if result else 1
		self.name = _FORMAT_MAP[main_type].format(next_num)

	def before_insert(self):
		if not self.system_date:
			self.system_date = now_datetime()
		if self.date:
			self.year = getdate(self.date).year
		if not self.created_by:
			self.created_by = frappe.session.user
		# Set default file_name only when no file is being uploaded at creation time.
		# If main_file is already set (upload at creation), after_save will set the
		# real filename after copying to the mount.
		if self.name and not self.main_file:
			self.file_name = self.name + ".pdf"

	def after_save(self):
		self._move_file_to_mount()
		self._sync_supporting_file_attachments()

	def _sync_supporting_file_attachments(self):
		"""Ensure every supporting file row that has a file_attachment URL also
		has a corresponding Frappe File record linked to the *parent* Cash Document
		(not the child row), so it appears in the sidebar attachment list."""
		if not self.supporting_files:
			return

		existing_urls = {
			r.file_url
			for r in frappe.get_all(
				"File",
				filters={
					"attached_to_doctype": "Cash Document",
					"attached_to_name": self.name,
				},
				fields=["file_url"],
			)
		}

		for row in self.supporting_files:
			url = row.get("file_attachment")
			if not url or url in existing_urls:
				continue
			fname = row.get("file_name") or url.rsplit("/", 1)[-1]
			frappe.db.sql(
				"""INSERT INTO `tabFile`
				    (name, file_name, file_url, attached_to_doctype, attached_to_name,
				     is_private, creation, modified, owner, modified_by, docstatus)
				   VALUES (%s, %s, %s, 'Cash Document', %s, 0,
				           NOW(), NOW(), %s, %s, 0)""",
				(
					frappe.generate_hash(length=10),
					fname, url, self.name,
					frappe.session.user, frappe.session.user,
				),
			)
			existing_urls.add(url)

	def _move_file_to_mount(self):
		"""If main_file points to a Frappe-local file, copy it to the shared
		network mount and update main_file / file_name to the /edox/ URL.

		Frappe uploads the file in a separate request before the document is
		saved, so at after_save time the File record's attached_to_name may
		not yet be committed.  We therefore resolve the disk path directly
		from the URL (via _url_to_path) instead of relying on a File-record
		lookup by attached_to_name.
		"""
		if not self.main_file:
			return
		# Already on the mount — nothing to do.
		if self.main_file.startswith("/edox/"):
			return
		# Only handle Frappe-managed file URLs.
		if not (self.main_file.startswith("/files/") or self.main_file.startswith("/private/files/")):
			return

		try:
			src_path = _url_to_path(self.main_file)
			if not src_path or not os.path.exists(src_path):
				frappe.log_error(
					f"Source not found: {src_path!r} (main_file={self.main_file!r})",
					"Cash Document — file move to mount failed",
				)
				return

			# Always store as {doc.name}.pdf on the mount to match legacy naming.
			filename = self.name + ".pdf"
			dest_dir = os.path.join(_MOUNT_BASE, self.name)
			dest_path = os.path.join(dest_dir, filename)
			os.makedirs(dest_dir, exist_ok=True)
			shutil.copy2(src_path, dest_path)

			original_url = self.main_file   # capture before we overwrite self.main_file
			edox_url = f"/edox/{self.name}/{filename}"
			frappe.db.set_value("Cash Document", self.name, {
				"main_file": edox_url,
				"file_name": filename,
			}, update_modified=False)
			self.main_file = edox_url
			self.file_name = filename

			# Update the Frappe File record URL so the sidebar points to /edox/.
			# Look up by file_url only — attached_to_name may not be committed yet.
			frappe.db.sql(
				"UPDATE `tabFile` SET file_url=%s, file_name=%s WHERE file_url=%s",
				(edox_url, filename, original_url),
			)
			# Remove the local copy — canonical location is now the mount.
			try:
				os.remove(src_path)
			except OSError:
				pass

		except Exception:
			frappe.log_error(frappe.get_traceback(), "Cash Document — file move to mount failed")

	def validate(self):
		# Prevent file attachment if main_type has not been chosen yet.
		# (The attachment widget saves immediately, before the user fills other fields.)
		if self.main_file and self.main_file.startswith(("/files/", "/private/files/")) and not self.main_type:
			frappe.throw(frappe._("Please set the Document Type before attaching a file."))

		if self.main_type:
			valid = _VALID_SUB_TYPES.get(self.main_type, set())
			sub = self.sub_type or ""

			if self.main_type in ("Cash", "Bank") and not sub:
				frappe.throw(
					frappe._("Sub Type is required for {0} documents.").format(self.main_type)
				)
			elif sub and sub not in valid:
				frappe.throw(
					frappe._("Sub Type '{0}' is not valid for {1} documents.").format(
						sub, self.main_type
					)
				)

		if self.date:
			self.year = getdate(self.date).year

		# Ensure file is copied to mount if main_file is set/changed
		self._move_file_to_mount()


# Whitelisted server functions

@frappe.whitelist()
def finalise(doc_name):
	"""Set status to final. Restricted to Cash Accountant / Cash Super User."""
	_require_role(["Cash Accountant", "Cash Super User", "Administrator"])
	doc = frappe.get_doc("Cash Document", doc_name)
	if doc.docstatus != 0:
		frappe.throw(frappe._("Finalise is only allowed on draft (pre-submit) documents."))
	if doc.status == "final":
		frappe.throw(frappe._("Document is already finalised."))
	frappe.db.set_value("Cash Document", doc_name, "status", "final")
	return "final"


@frappe.whitelist()
def _create_journal_entry(doc):
	"""Validate JE data and post a submitted ERPNext Journal Entry for a Cash Document.

	Called from finalise2(). Resolves GL accounts via Cash GL Account Mapping,
	handles multi-currency, and writes the JE name back to doc.je_ref.
	"""
	# ── 1. Pre-condition guards ───────────────────────────────────────────
	if doc.company == "Unknown":
		frappe.throw(frappe._("Company is 'Unknown' — resolve the company before finalising."))

	if doc.je_ref:
		frappe.throw(
			frappe._("A Journal Entry ({0}) is already linked to this document.").format(doc.je_ref)
		)

	if not doc.jeid:
		frappe.throw(
			frappe._("Transaction Details not imported (JE ID is missing). Import JE details first.")
		)

	for fname, label in (("account_name", "Account Name"), ("contra_account_name", "Contra Account Name"), ("je_currency", "Currency")):
		if not (doc.get(fname) or "").strip():
			frappe.throw(
				frappe._("{0} is missing. Import JE details before finalising.").format(label)
			)

	txn_amount = (doc.main_debit or 0) + (doc.main_credit or 0)
	if txn_amount <= 0:
		frappe.throw(frappe._("Transaction amount is zero. Import JE details before finalising."))

	# ── 2. Resolve GL accounts ────────────────────────────────────────────
	def _resolve(third_party_name):
		account = frappe.db.get_value(
			"Cash GL Account Mapping",
			{"third_party_name": third_party_name, "cash_company": doc.company},
			"erpnext_account",
		)
		if not account:
			frappe.throw(
				frappe._(
					"No GL account mapping found for '{0}' (Company: {1}). "
					"Add it in Cash GL Account Mapping before finalising."
				).format(third_party_name, doc.company)
			)
		return account

	debit_account  = _resolve(doc.account_name)
	credit_account = _resolve(doc.contra_account_name)

	# ── 3. Derive ERPNext company from the mapped account ─────────────────
	erpnext_company = frappe.db.get_value("Account", debit_account, "company")
	credit_company  = frappe.db.get_value("Account", credit_account, "company")
	if erpnext_company != credit_company:
		frappe.throw(
			frappe._(
				"GL mapping error: '{0}' belongs to {1} but '{2}' belongs to {3}. "
				"Both accounts must be in the same company."
			).format(debit_account, erpnext_company, credit_account, credit_company)
		)

	company_currency  = frappe.db.get_value("Company", erpnext_company, "default_currency")
	is_multi_currency = (doc.je_currency or "").strip().upper() != (company_currency or "").strip().upper()

	# ── 4. Determine debit / credit sides and amounts ─────────────────────
	# main_debit/credit is in je_currency; sec_debit/credit is in company base currency.
	# If both are set (shouldn't happen), main_debit takes precedence.
	if (doc.main_debit or 0) > 0:
		debit_side  = debit_account
		credit_side = credit_account
		txn_amt  = doc.main_debit
		base_amt = doc.sec_debit or doc.main_debit
	else:
		debit_side  = credit_account
		credit_side = debit_account
		txn_amt  = doc.main_credit
		base_amt = doc.sec_credit or doc.main_credit

	exchange_rate = round(base_amt / txn_amt, 9) if txn_amt else 1.0
	cost_center   = frappe.db.get_value("Company", erpnext_company, "cost_center")

	# ── 5. Build and submit the Journal Entry ─────────────────────────────
	je = frappe.get_doc({
		"doctype":       "Journal Entry",
		"voucher_type":  "Journal Entry",
		"company":       erpnext_company,
		"posting_date":  doc.je_doc_date or doc.date,
		"cheque_no":     doc.jeid,
		"cheque_date":   doc.je_doc_date or doc.date,
		"user_remark":   ((doc.je_details or "") + " [Cash Doc: " + doc.name + "]").strip(),
		"multi_currency": 1 if is_multi_currency else 0,
		"accounts": [
			{
				"account":                      debit_side,
				"account_currency":             doc.je_currency,
				"exchange_rate":                exchange_rate,
				"debit_in_account_currency":    txn_amt,
				"debit":                        base_amt,
				"credit_in_account_currency":   0,
				"credit":                       0,
				"cost_center":                  cost_center,
			},
			{
				"account":                      credit_side,
				"account_currency":             doc.je_currency,
				"exchange_rate":                exchange_rate,
				"debit_in_account_currency":    0,
				"debit":                        0,
				"credit_in_account_currency":   txn_amt,
				"credit":                       base_amt,
				"cost_center":                  cost_center,
			},
		],
	})
	je.insert(ignore_permissions=True)
	je.submit()

	# ── 6. Write JE name back to Cash Document ────────────────────────────
	frappe.db.set_value("Cash Document", doc.name, "je_ref", je.name, update_modified=False)
	return je.name


def finalise2(doc_name):
	"""Set final_status2 to final2. Restricted to Cash Checker / Cash Super User.

	Requires Status 1 to already be 'final' — enforces the two-step approval sequence.
	Validates and posts an ERPNext Journal Entry before marking the document final.
	"""
	_require_role(["Cash Checker", "Cash Super User", "Administrator"])
	doc = frappe.get_doc("Cash Document", doc_name)
	if doc.docstatus != 0:
		frappe.throw(frappe._("Finalise 2 is only allowed on draft (pre-submit) documents."))
	if doc.status != "final":
		frappe.throw(frappe._("Status 1 must be finalised before Status 2 can be set."))
	if doc.final_status2 == "final2":
		frappe.throw(frappe._("Document is already finalised (Status 2)."))

	_create_journal_entry(doc)  # validates data completeness and posts the JE

	frappe.db.set_value("Cash Document", doc_name, "final_status2", "final2")
	return "final2"


@frappe.whitelist()
def unfinalise(doc_name):
	"""Reset status back to pending. Restricted to Cash Super User.

	Blocked if Status 2 is already final — must unfinalise Status 2 first.
	"""
	_require_role(["Cash Super User", "Administrator"])
	doc = frappe.get_doc("Cash Document", doc_name)
	if doc.status != "final":
		frappe.throw(frappe._("Document is not finalised."))
	if doc.final_status2 == "final2":
		frappe.throw(frappe._("Status 2 is already final — unfinalise Status 2 first."))
	frappe.db.set_value("Cash Document", doc_name, "status", "pending")
	return "pending"


@frappe.whitelist()
def unfinalise2(doc_name):
	"""Reset final_status2 back to pending2 and cancel the linked Journal Entry.

	Restricted to Cash Super User.
	"""
	_require_role(["Cash Super User", "Administrator"])
	doc = frappe.get_doc("Cash Document", doc_name)
	if doc.final_status2 != "final2":
		frappe.throw(frappe._("Status 2 is not finalised."))

	# Cancel the linked JE before resetting the status
	if doc.je_ref:
		je = frappe.get_doc("Journal Entry", doc.je_ref)
		if je.docstatus == 1:
			je.cancel()
		frappe.db.set_value("Cash Document", doc_name, "je_ref", None, update_modified=False)

	frappe.db.set_value("Cash Document", doc_name, "final_status2", "pending2")
	return "pending2"


@frappe.whitelist()
def add_flag(doc_name, flag_type, comment):
	"""Append a review flag row to document_flags."""
	comment = (comment or "").strip()
	if not comment:
		frappe.throw(frappe._("Flag comment cannot be empty."))

	valid_types = {
		"Review Required", "Approved", "Rejected",
		"Query", "Hold", "Priority", "Revision Needed",
	}
	if flag_type not in valid_types:
		frappe.throw(frappe._("Invalid flag type: {0}").format(flag_type))

	doc = frappe.get_doc("Cash Document", doc_name)
	doc.append("document_flags", {
		"flag_type":      flag_type,
		"flag_date":      now_datetime(),
		"flagged_by":     frappe.session.user,
		"flagged_by_role": _primary_cash_role(),
		"comments":       comment,
	})
	doc.save(ignore_permissions=True)
	return len(doc.document_flags)


@frappe.whitelist()
def clear_flags(doc_name):
	"""Remove all review flags. Restricted to Cash Super User."""
	_require_role(["Cash Super User", "Administrator"])
	frappe.db.delete("Cash Document Flag", {"parent": doc_name})
	return 0


@frappe.whitelist()
def resync_counters():
	"""
	Resync Cash Voucher Series counters to MAX(numeric suffix) across all
	existing Cash Document records.  Call this after any bulk import of
	legacy data to prevent numbering collisions.

	Restricted to Cash Super User / Administrator.
	"""
	_require_role(["Cash Super User", "Administrator"])
	from kgk_customisations.patches.v1_0.resync_cash_voucher_series import (
		resync_counters as _resync,
	)
	return _resync()


@frappe.whitelist()
def import_je_details(file_url):
	"""Read an uploaded XLS/XLSX Fantasy Export file and populate Transaction
	Details fields on all Cash Documents. Matching priority:
	  1. JEID (col 3 in file) matched against Cash Document jeid field.
	  2. Invoice # (col 17 in file) matched against Cash Document file_name
	     (extension stripped), for legacy records where JEID is absent.
	  3. migration_reference matched against JEID first, then Invoice #
	     (controlled fallback for migrated legacy records).
	  4. If all keys are absent/unmatched, the record is skipped.
	The uploaded Frappe File doc is deleted after processing.

	Large-volume safeguards:
	  - openpyxl read_only streaming for .xlsx — workbook is never fully in memory
	  - xlrd datemode captured once (not per cell) to avoid O(n) file re-opens
	  - UPDATEs committed in chunks of CHUNK_SIZE to cap transaction size and
	    allow partial saves if the worker times out
	  - not_found list capped in the HTTP response; full list written to Error Log
	"""
	CHUNK_SIZE = 500  # commit every N matched rows

	# Resolve the uploaded file
	file_docs = frappe.get_all(
		"File",
		filters={"file_url": file_url},
		fields=["name", "file_name"],
		limit=1,
	)
	if not file_docs:
		frappe.throw(frappe._("Uploaded file not found: {0}").format(file_url))

	file_doc = frappe.get_doc("File", file_docs[0].name)
	ext = os.path.splitext(file_doc.file_name)[1].lower()

	if ext not in (".xls", ".xlsx"):
		frappe.throw(frappe._("Only .xls and .xlsx files are supported."))

	disk_path = file_doc.get_full_path()

	# Capture xlrd datemode once so _to_date doesn't re-open the file per cell
	_xlrd_datemode = None
	if ext == ".xls":
		try:
			import xlrd as _xlrd
			_xlrd_datemode = _xlrd.open_workbook(disk_path).datemode
		except ImportError:
			frappe.throw(frappe._("xlrd is required for .xls files. Install with: pip install xlrd"))

	def _iter_rows():
		"""Yield raw row value lists from the spreadsheet, skipping the header."""
		if ext == ".xlsx":
			import openpyxl
			wb = openpyxl.load_workbook(disk_path, read_only=True, data_only=True)
			sh = wb.active
			first = True
			for row in sh.iter_rows(values_only=True):
				if first:
					first = False
					continue
				yield list(row)
			wb.close()
		else:
			import xlrd
			wb = xlrd.open_workbook(disk_path)
			sh = wb.sheets()[0]
			for i in range(1, sh.nrows):
				yield sh.row_values(i)

	def _to_date(val):
		"""Convert a cell value to an ISO date string regardless of source format."""
		if val is None or val == "":
			return None
		if hasattr(val, "isoformat"):  # openpyxl returns date/datetime objects
			return val.date().isoformat() if hasattr(val, "date") else val.isoformat()
		if isinstance(val, str):
			return val.strip() or None
		try:  # xlrd float serial — use pre-captured datemode, never re-open the file
			import xlrd
			return xlrd.xldate_as_datetime(val, _xlrd_datemode).date().isoformat()
		except Exception:
			return None

	# Build lookup indices — one entry per key (first data row wins).
	# jeid_index   : keyed by JEID (col 3) for rows where JEID is present.
	# invoice_index: keyed by Invoice # (col 17) for ANY row that has one,
	#                including rows that also have a JEID. This allows legacy
	#                Cash Documents (no jeid set) to be matched by file name,
	#                then have the JEID written back from the XLS row.
	#                Each entry carries '_jeid' so it can be saved on the doc.
	# Group header rows (col 0 non-empty) are skipped in both cases.
	# Cols 12 and 15 are running totals (derived) and are not stored.
	jeid_index    = {}
	invoice_index = {}
	for row in _iter_rows():
		if row[0]:  # group header — skip
			continue
		jeid        = str(row[3]).strip()  if row[3]                     else ""
		invoice_num = str(row[17]).strip() if len(row) > 17 and row[17] else ""
		fields = {
			"account_id":          str(int(row[1])) if row[1] else "",
			"contra_account_id":   str(int(row[2])) if row[2] else "",
			"je_doc_date":         _to_date(row[4]),
			"je_line_date":        _to_date(row[5]),
			"account_name":        str(row[6]).strip() if row[6] else "",
			"contra_account_name": str(row[7]).strip() if row[7] else "",
			"je_details":          str(row[8]).strip() if row[8] else "",
			"je_currency":         str(row[9]).strip() if row[9] else "",
			"main_debit":          row[10] or 0,
			"main_credit":         row[11] or 0,
			"sec_debit":           row[13] or 0,
			"sec_credit":          row[14] or 0,
			"je_audit":            str(row[16]).strip() if row[16] else "",
			"je_supplier":         str(row[18]).strip() if len(row) > 18 and row[18] else "",
			"je_user":             str(row[19]).strip() if len(row) > 19 and row[19] else "",
		}
		if jeid:
			if jeid not in jeid_index:
				jeid_index[jeid] = fields
		if invoice_num:
			if invoice_num not in invoice_index:
				# Store the JEID alongside so it can be written back to the doc
				invoice_index[invoice_num] = {**fields, "_jeid": jeid}
		if not jeid and not invoice_num:
			pass  # both absent — row is unidentifiable, skip

	# Find all Cash Documents that have at least one lookup key to match against
	docs = frappe.db.sql(
		"""SELECT name, jeid, file_name, migration_reference FROM `tabCash Document`
		   WHERE (jeid IS NOT NULL AND jeid != '')
		      OR (file_name IS NOT NULL AND file_name != '')
		      OR (migration_reference IS NOT NULL AND migration_reference != '')""",
		as_dict=True,
	)

	matched = []
	not_found = []
	chunk_count = 0

	for doc in docs:
		jeid = str(doc["jeid"]).strip() if doc.get("jeid") else ""
		file_name = str(doc["file_name"]).strip() if doc.get("file_name") else ""
		migration_ref = str(doc["migration_reference"]).strip() if doc.get("migration_reference") else ""
		recovered_jeid = ""

		# Priority 1: match by JEID
		if jeid and jeid in jeid_index:
			row = jeid_index[jeid]
			recovered_jeid = jeid
		# Priority 2: match Invoice # against file_name (extension stripped)
		elif file_name:
			lookup_key = os.path.splitext(file_name)[0]
			if lookup_key in invoice_index:
				row = invoice_index[lookup_key]
				recovered_jeid = row.get("_jeid") or ""
			else:
				# Priority 3: migration_reference fallback (first JEID, then Invoice #)
				if migration_ref and migration_ref in jeid_index:
					row = jeid_index[migration_ref]
					recovered_jeid = migration_ref
				elif migration_ref and migration_ref in invoice_index:
					row = invoice_index[migration_ref]
					recovered_jeid = row.get("_jeid") or ""
				else:
					not_found.append(doc["name"])
					continue
		# Priority 3 fallback when file_name is unavailable
		elif migration_ref:
			if migration_ref in jeid_index:
				row = jeid_index[migration_ref]
				recovered_jeid = migration_ref
			elif migration_ref in invoice_index:
				row = invoice_index[migration_ref]
				recovered_jeid = row.get("_jeid") or ""
			else:
				not_found.append(doc["name"])
				continue
		else:
			not_found.append(doc["name"])
			continue
		frappe.db.sql(
			"""UPDATE `tabCash Document` SET
				account_id          = %(account_id)s,
				contra_account_id   = %(contra_account_id)s,
				je_doc_date         = %(je_doc_date)s,
				je_line_date        = %(je_line_date)s,
				account_name        = %(account_name)s,
				contra_account_name = %(contra_account_name)s,
				je_details          = %(je_details)s,
				je_currency         = %(je_currency)s,
				main_debit          = %(main_debit)s,
				main_credit         = %(main_credit)s,
				sec_debit           = %(sec_debit)s,
				sec_credit          = %(sec_credit)s,
				je_audit            = %(je_audit)s,
				je_supplier         = %(je_supplier)s,
				je_user             = %(je_user)s
			WHERE name = %(name)s""",
			{**row, "name": doc["name"]},
		)
		# Backfill JEID when the document JEID is blank and XLS provides one.
		if not jeid and recovered_jeid:
			frappe.db.sql(
				"UPDATE `tabCash Document` SET jeid = %(jeid)s WHERE name = %(name)s",
				{"jeid": recovered_jeid, "name": doc["name"]},
			)
		matched.append(doc["name"])
		chunk_count += 1
		if chunk_count % CHUNK_SIZE == 0:
			frappe.db.commit()  # commit each chunk — caps transaction size

	if chunk_count % CHUNK_SIZE != 0:
		frappe.db.commit()  # commit the final partial chunk

	# Write not_found to Error Log if large, so the HTTP response stays small
	NOT_FOUND_RESPONSE_LIMIT = 50
	if len(not_found) > NOT_FOUND_RESPONSE_LIMIT:
		frappe.log_error(
			title="JE Import — unmatched Cash Documents",
			message="\n".join(not_found),
		)

	# Delete the uploaded file only after data is safely committed
	file_doc.delete(ignore_permissions=True)
	frappe.db.commit()

	return {
		"matched":         len(matched),
		"not_found":       not_found[:NOT_FOUND_RESPONSE_LIMIT],
		"not_found_total": len(not_found),
		"xls_jeid_rows":   len(jeid_index),
		"xls_invoice_rows": len(invoice_index),
	}


@frappe.whitelist()
def download_merged_pdf(doc_name):
	"""Generate a merged PDF: cover sheet (print format) + main file + supporting files.

	Returns the PDF as a file download response.
	Steps:
	  1. Render 'Cash Document Print' → PDF bytes via frappe.get_print (server-side HTML, no network needed)
	  2. Read the main PDF from the e-dox mount (if set)
	  3. Read each supporting _A/_B/… PDF from the mount
	  4. Merge all into one PDF with pypdf
	  5. Stream back as a file download
	"""
	from pypdf import PdfWriter, PdfReader

	doc = frappe.get_doc("Cash Document", doc_name)

	# --- 1. Cover sheet PDF via frappe.get_print ---
	cover_bytes = frappe.get_print(
		doctype="Cash Document",
		name=doc_name,
		print_format="Cash Document Print",
		as_pdf=True,
		no_letterhead=1,
	)

	writer = PdfWriter()
	skipped = []

	def _append_file(path, label):
		"""Append all pages from a PDF file on disk into the writer.

		Read the entire file into a BytesIO buffer first so that the
		PdfReader always has access to the stream even after the OS file
		handle is closed — pypdf defers content-stream reads until
		writer.write(), so a closed file handle causes silent data loss.
		"""
		try:
			with open(path, "rb") as fh:
				buf = io.BytesIO(fh.read())          # fully buffer in memory
			reader = PdfReader(buf)                  # reader holds BytesIO, stays open
			for page in reader.pages:
				writer.add_page(page)
		except FileNotFoundError:
			skipped.append(f"{label}: file not found at {path}")
		except Exception as exc:
			skipped.append(f"{label}: {exc}")

	# Cover sheet (in-memory bytes — already a BytesIO-compatible source)
	try:
		reader = PdfReader(io.BytesIO(cover_bytes))
		for page in reader.pages:
			writer.add_page(page)
	except Exception as exc:
		frappe.throw(frappe._("Could not read generated cover sheet: {0}").format(exc))

	# --- 2. Main file ---
	if doc.main_file:
		main_path = _url_to_path(doc.main_file)
		if main_path:
			_append_file(main_path, f"main file ({doc.main_file})")
		else:
			skipped.append(f"main file: unrecognised URL scheme ({doc.main_file!r})")

	# --- 3. Supporting files ---
	for row in doc.supporting_files:
		url = row.file_attachment or ""
		if not url:
			continue
		supp_path = _url_to_path(url)
		if supp_path:
			_append_file(supp_path, f"supporting {row.file_suffix or ''} ({url})")
		else:
			skipped.append(f"supporting {row.file_suffix or ''}: unrecognised URL scheme ({url!r})")

	# Log any skipped files for the user to see (non-fatal)
	if skipped:
		frappe.log_error("\n".join(skipped), f"download_merged_pdf: {doc_name} — skipped files")
		frappe.msgprint(
			frappe._("Some attachments could not be included (logged to Error Log):<br>")
			+ "<br>".join(skipped),
			alert=True,
			indicator="orange",
		)

	# --- 4. Stream result ---
	output_buf = io.BytesIO()
	writer.write(output_buf)
	pdf_bytes = output_buf.getvalue()

	# Log page count for diagnostics
	frappe.log_error(
		f"pages={len(writer.pages)}, size={len(pdf_bytes)}, skipped={skipped}",
		f"download_merged_pdf OK: {doc_name}",
	)

	# Use type="pdf" to match Frappe's own print download path (as_pdf handler)
	frappe.local.response.filename = f"{doc_name}.pdf"
	frappe.local.response.filecontent = pdf_bytes
	frappe.local.response.type = "pdf"


# Internal helpers

def _require_role(roles):
	user_roles = set(frappe.get_roles(frappe.session.user))
	if not user_roles.intersection(roles):
		frappe.throw(
			frappe._("You do not have permission to perform this action."),
			frappe.PermissionError,
		)


def _primary_cash_role():
	"""Return the highest Cash role the current user holds, for audit logging."""
	priority = ["Cash Super User", "Cash Accountant", "Cash Checker", "Cash Basic User"]
	user_roles = set(frappe.get_roles(frappe.session.user))
	for role in priority:
		if role in user_roles:
			return role
	return "User"
