# Copyright (c) 2025, Apjakal IT Solutions and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import json
import os
import re
from kgk_customisations.utils.ocr_utils import extract_ocr_fields_from_text, get_consolidated_ocr_data
from kgk_customisations.utils.excel_utils import create_styled_excel_workbook, create_download_response
from kgk_customisations.kgk_customisations.utils.permission_manager import PermissionManager


# Extraction logic moved to utils/ocr_utils.py for reusability


class OCRDataUpload(Document):
	@frappe.whitelist()
	def preview_excel_data(self):
		"""Load Excel data into child table for preview"""
		try:
			print(f"Preview method called - file_upload: {self.file_upload}")
			
			if not self.file_upload:
				error_msg = "Please upload an Excel file first"
				frappe.msgprint(error_msg, indicator="red")
				return {"success": False, "message": error_msg}

			# Mark file as private for security
			try:
				file_doc = frappe.get_doc("File", {"file_url": self.file_upload})
				if not file_doc.is_private:
					file_doc.is_private = 1
					# Save with permission check - system operation (security enhancement)
					PermissionManager.save_with_permission_check(file_doc, ignore_for_system=True)
					frappe.db.commit()
			except Exception as e:
				error_msg = f"Error accessing file: {str(e)}"
				print(f"File access error: {error_msg}")
				frappe.msgprint(error_msg, indicator="red")
				return {"success": False, "message": error_msg}

			# Get file path
			if file_doc.file_url.startswith('/private/'):
				file_path = frappe.get_site_path("private", "files", file_doc.file_name)
			else:
				file_path = frappe.get_site_path("public", "files", file_doc.file_name)

			print(f"File path: {file_path}")

			# Map expected Excel columns to DocType fields (define this first)
			# These should match the actual Excel file column headers
			excel_to_field_mapping = {
				# Common field variations - update these based on your actual Excel headers
				"Sequence": "sequence",
				"Created On": "created_on",				
				"Batch Name": "batch_name",
				"Text Data": "text_data",
				
				# Lot ID variations
				"Lot ID 1": "lot_id_1",
				"Lot ID 2": "lot_id_2",
				"Sub Lot ID": "sub_lot_id",
				
				# Result/analysis fields
				"Result": "result",
				"Color": "color",
				"Blue UV": "blue_uv",
				"Brown": "brown",
				"Yellow UV": "yellow_uv",
				"Type": "type",
				"Fancy Yellow": "fancy_yellow",
			}

			# Read Excel file
			import os
			if not os.path.exists(file_path):
				error_msg = f"File not found at expected location: {file_path}"
				print(f"File not found error: {error_msg}")
				frappe.msgprint(error_msg, indicator="red")
				return {"success": False, "message": error_msg}

			try:
				import pandas as pd
				df = pd.read_excel(file_path)
				print(f"Excel file loaded successfully. Rows: {len(df)}, Columns: {list(df.columns)}")
				
				# Log first few rows to understand data structure
				if len(df) > 0:
					print(f"Sample data from first row:")
					for col in df.columns:
						sample_value = df.iloc[0][col] if not pd.isna(df.iloc[0][col]) else "N/A"
						print(f"  '{col}': {sample_value}")
				
				# Check which Excel columns will be mapped to which fields
				print(f"\nColumn mapping analysis:")
				found_mappings = []
				missing_columns = []
				
				for excel_col, field_name in excel_to_field_mapping.items():
					if excel_col in df.columns:
						found_mappings.append(f"  FOUND: '{excel_col}' -> {field_name}")
						print(f"  FOUND: '{excel_col}' -> {field_name}")
					else:
						missing_columns.append(excel_col)
				
				print(f"\nUnmapped columns in Excel file:")
				for col in df.columns:
					if col not in excel_to_field_mapping:
						print(f"  UNMAPPED: '{col}'")
				
				if missing_columns:
					print(f"\nMissing expected columns: {missing_columns}")
						
			except Exception as e:
				error_msg = f"Error reading Excel file: {str(e)}"
				print(f"Excel read error: {error_msg}")
				frappe.msgprint(error_msg, indicator="red")
				return {"success": False, "message": error_msg}

			# Clear existing items
			self.set('items', [])

			# Add items to child table
			rows_processed = 0
			for index, row in df.iterrows():
				try:
					child = self.append('items', {})
					populated_fields = []
					
					# First, map direct Excel columns to fields
					for excel_col, field_name in excel_to_field_mapping.items():
						if excel_col in df.columns:
							value = row[excel_col]
							if pd.notna(value):
								# Handle date fields specially
								if field_name == 'created_on':
									try:
										# Try to parse different date formats
										if isinstance(value, str):
											# Try DD/MM/YYYY format first
											from datetime import datetime
											try:
												date_obj = datetime.strptime(value, '%d/%m/%Y').date()
											except ValueError:
												try:
													# Try MM/DD/YYYY format
													date_obj = datetime.strptime(value, '%m/%d/%Y').date()
												except ValueError:
													try:
														# Try YYYY-MM-DD format
														date_obj = datetime.strptime(value, '%Y-%m-%d').date()
													except ValueError:
														# Default to today if parsing fails
														date_obj = datetime.today().date()
											setattr(child, field_name, date_obj)
											populated_fields.append(f"{field_name}={date_obj}")
										elif hasattr(value, 'date'):
											# It's already a datetime object
											setattr(child, field_name, value.date())
											populated_fields.append(f"{field_name}={value.date()}")
										else:
											# Default to today
											from datetime import date
											setattr(child, field_name, date.today())
											populated_fields.append(f"{field_name}={date.today()}")
									except Exception as e:
										print(f"Error parsing date {value}: {str(e)}")
										from datetime import date
										setattr(child, field_name, date.today())
										populated_fields.append(f"{field_name}={date.today()}(default)")
								else:
									# Convert to string and truncate if needed
									str_value = str(value)
									if field_name == 'text_data':
										# OCR output can be very long, limit to 8000 chars
										str_value = str_value[:8000]
									elif field_name == 'image_data':
										# Image data can be very long, limit to 10000 chars
										str_value = str_value[:10000]
									else:
										# Other fields, limit to reasonable length
										str_value = str_value[:200]
									setattr(child, field_name, str_value)
									populated_fields.append(f"{field_name}={str_value[:50]}...")
							else:
								# Value is NaN/empty
								populated_fields.append(f"{field_name}=EMPTY")
						else:
							# Excel column not found
							populated_fields.append(f"{excel_col}->NOT_FOUND")
					
					# Now, if we have OCR output text, extract refined values using our extraction logic
					if hasattr(child, 'text_data') and child.text_data:
						try:
							extracted_fields = extract_ocr_fields_from_text(child.text_data)
							
							# Apply fallback logic from original Excel values (like in reference code)
							# Only update fields if they're not already populated from Excel, with fallbacks
							
							# Result field
							if not getattr(child, 'result', None) and extracted_fields.get('Result'):
								child.result = extracted_fields['Result']
								populated_fields.append(f"result(extracted)={extracted_fields['Result']}")
							
							# Color field with fallback logic
							if not getattr(child, 'color', None):
								if extracted_fields.get('Color'):
									child.color = extracted_fields['Color']
									populated_fields.append(f"color(extracted)={extracted_fields['Color']}")
								else:
									# Fallback: check if original Excel had valid color
									original_color = str(getattr(child, 'color', '') or '').strip().upper()
									if re.fullmatch(r"[DEFGHIJK][+\-]?|\d{3}", original_color):
										child.color = original_color
										populated_fields.append(f"color(fallback)={original_color}")
							
							# Blue UV field
							if not getattr(child, 'blue_uv', None) and extracted_fields.get('Blue UV'):
								child.blue_uv = extracted_fields['Blue UV']
								populated_fields.append(f"blue_uv(extracted)={extracted_fields['Blue UV']}")
							
							# Brown field with specific fallback logic
							if not getattr(child, 'brown', None):
								if extracted_fields.get('Brown'):
									child.brown = extracted_fields['Brown']
									populated_fields.append(f"brown(extracted)={extracted_fields['Brown']}")
								else:
									# Check if original Excel had "NOT MEASURED"
									original_brown = str(getattr(child, 'brown', '') or '').strip().upper()
									if original_brown == "NOT MEASURED":
										child.brown = "NOT MEASURED"
										populated_fields.append(f"brown(fallback)=NOT MEASURED")
							
							# Yellow UV field with special logic
							if not getattr(child, 'yellow_uv', None):
								if extracted_fields.get('Yellow UV'):
									child.yellow_uv = extracted_fields['Yellow UV']
									populated_fields.append(f"yellow_uv(extracted)={extracted_fields['Yellow UV']}")
								elif "YELL UV" in child.text_data.upper():
									child.yellow_uv = "YELL UV"
									populated_fields.append(f"yellow_uv(detected)=YELL UV")
							
							# Type field with fallback
							if not getattr(child, 'type', None):
								if extracted_fields.get('Type'):
									child.type = extracted_fields['Type']
									populated_fields.append(f"type(extracted)={extracted_fields['Type']}")
								else:
									# Check if original Excel had valid type
									original_type = str(getattr(child, 'type', '') or '').strip().upper()
									if original_type in {"MIXED", "WHITE", "BLUE OR GRAY", "GRAY", "BROWN"}:
										child.type = original_type
										populated_fields.append(f"type(fallback)={original_type}")
							
							# Fancy Yellow field
							if not getattr(child, 'fancy_yellow', None) and extracted_fields.get('Fancy Yellow'):
								child.fancy_yellow = extracted_fields['Fancy Yellow']
								populated_fields.append(f"fancy_yellow(extracted)={extracted_fields['Fancy Yellow']}")
									
						except Exception as e:
							print(f"Error extracting from OCR text for row {index}: {str(e)}")
							import traceback
							traceback.print_exc()
					else:
						# No OCR data available
						populated_fields.append("no_ocr_data")
					
					# Set created_on to today if not set
					if not getattr(child, 'created_on', None):
						from datetime import date
						child.created_on = date.today()
						populated_fields.append(f"created_on(default)={date.today()}")
					
					# Log what was populated for first few rows
					if index < 5:
						print(f"Row {index} populated fields: {', '.join(populated_fields)}")
					
					rows_processed += 1
				except Exception as e:
					print(f"Error processing row {index}: {str(e)}")
					import traceback
					traceback.print_exc()
					continue

			# Save the document
			try:
				self.save()
				frappe.db.commit()
			except Exception as e:
				error_msg = f"Error saving document: {str(e)}"
				print(f"Save error: {error_msg}")
				frappe.msgprint(error_msg, indicator="red")
				return {"success": False, "message": error_msg}

			msg = f"Successfully loaded {rows_processed} rows from Excel file."
			# update "total_records" with rows_processed
			self.total_records = rows_processed
			self.save()
			frappe.db.commit()
			frappe.msgprint(msg, indicator="green")
			print(f"Success: {msg}")

			return {
				"success": True,
				"message": msg,
				"rows_loaded": rows_processed
			}

		except Exception as e:
			error_msg = f"Unexpected error in preview_excel_data: {str(e)}"
			frappe.msgprint(error_msg, indicator="red")
			print(f"Unexpected error: {error_msg}")
			import traceback
			traceback.print_exc()
			return {"success": False, "message": error_msg}


@frappe.whitelist()
def download_cumulative_report():
	"""Generate and download cumulative report with refined columns - using centralized utilities"""
	try:
		# Get consolidated OCR data using centralized function
		ocr_data = get_consolidated_ocr_data(include_refined=True, format="dict")
		
		if not ocr_data or len(ocr_data) == 0:
			return {"success": False, "message": "No OCR data found for report generation"}
		
		total_records = len(ocr_data)
		
		# For very large datasets, use background job
		if total_records > 5000:
			job = frappe.enqueue(
				method='kgk_customisations.kgk_customisations.doctype.ocr_data_upload.ocr_data_upload.generate_large_report_background',
				timeout=1800,
				is_async=True,
				job_name=f"ocr_report_{frappe.session.user}_{frappe.utils.now()}"
			)
			
			return {
				"success": True,
				"message": f"Report generation started in background for {total_records} records. You will receive a notification when it's ready.",
				"is_background": True,
				"job_id": job.id,
				"records_count": total_records
			}
		
		# For smaller datasets, process immediately
		return generate_small_report_immediate_centralized(ocr_data)
		
	except Exception as e:
		error_msg = f"Error in download_cumulative_report: {str(e)}"
		frappe.log_error(error_msg)
		return {"success": False, "message": error_msg}
def generate_small_report_immediate_centralized(ocr_data):
	"""Generate report immediately using centralized utilities for smaller datasets"""
	try:
		from kgk_customisations.utils.excel_utils import create_download_response
		from openpyxl import Workbook
		from openpyxl.styles import PatternFill, Font
		from openpyxl.utils import get_column_letter
		
		# Create Excel workbook
		wb = Workbook()
		ws = wb.active
		ws.title = "Cumulative OCR Report"
		
		# Define headers including refined columns
		headers = [
			"Upload Date", "Sequence", "Created On", "Batch Name", "Text Data",
			"Lot ID 1", "Lot ID 2", "Sub Lot ID", "Result", "Color", 
			"Blue UV", "Yellow UV", "Brown", "Type", "Fancy Yellow",
			"Refined Result", "Refined Color", "Refined Blue UV", "Refined Brown",
			"Refined Yellow UV", "Refined Type", "Refined Fancy Yellow"
		]
		ws.append(headers)
		
		# Process each record
		processed_count = 0
		for item in ocr_data:
			# Base row data
			row = [
				item.get("upload_date", ""),
				item.get("sequence", ""),
				item.get("created_on", ""),
				item.get("batch_name", ""),
				(item.get("text_data", "") or "")[:500],  # Truncate for display
				item.get("lot_id_1", ""),
				item.get("lot_id_2", ""),
				item.get("sub_lot_id", ""),
				item.get("result", ""),
				item.get("color", ""),
				item.get("blue_uv", ""),
				item.get("yellow_uv", ""),
				item.get("brown", ""),
				item.get("type", ""),
				item.get("fancy_yellow", "")
			]
			
			# Add refined columns (these should already be in the data from get_consolidated_ocr_data)
			row.extend([
				item.get("refined_result", ""),
				item.get("refined_color", ""),
				item.get("refined_blue_uv", ""),
				item.get("refined_brown", ""),
				item.get("refined_yellow_uv", ""),
				item.get("refined_type", ""),
				item.get("refined_fancy_yellow", "")
			])
			
			ws.append(row)
			processed_count += 1
		
		# Apply formatting using centralized utility
		from kgk_customisations.utils.excel_utils import apply_standard_header_formatting
		apply_standard_header_formatting(ws, 1)
		
		# Highlight refined columns manually since our utils might not handle this exactly
		yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
		yellow_header_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
		header_font = Font(bold=True)
		
		# Refined columns start at column 16 (P) and go to column 22 (V)
		refined_start_col = 16  # "Refined Result" column
		refined_end_col = 22    # "Refined Fancy Yellow" column
		
		# Apply yellow highlighting to refined column headers
		for col_idx in range(refined_start_col, refined_end_col + 1):
			cell = ws.cell(row=1, column=col_idx)
			cell.fill = yellow_header_fill
			cell.font = header_font
		
		# Apply yellow highlighting to all refined column data cells
		for row_idx in range(2, ws.max_row + 1):
			for col_idx in range(refined_start_col, refined_end_col + 1):
				cell = ws.cell(row=row_idx, column=col_idx)
				cell.fill = yellow_fill
		
		# Set column widths
		for col_idx in range(1, len(headers) + 1):
			column_letter = get_column_letter(col_idx)
			if col_idx == 5:  # Text Data column
				ws.column_dimensions[column_letter].width = 30
			elif col_idx >= refined_start_col:  # Refined columns
				ws.column_dimensions[column_letter].width = 12
			else:
				ws.column_dimensions[column_letter].width = 15
		
		# Create download using centralized function
		filename = f"cumulative_ocr_report_{frappe.utils.today().replace('-', '_')}.xlsx"
		download_result = create_download_response(wb, filename)
		
		if download_result.get("success"):
			return {
				"success": True,
				"message": f"Report generated with {processed_count} records including refined analysis",
				"download_triggered": True,
				"records_count": processed_count
			}
		else:
			return {"success": False, "message": f"Error creating download: {download_result.get('error')}"}
		
	except Exception as e:
		error_msg = f"Error generating immediate report with centralized utilities: {str(e)}"
		frappe.log_error(error_msg)
		import traceback
		traceback.print_exc()
		return {"success": False, "message": error_msg}


def generate_report_using_centralized_functions(ocr_data):
	"""Generate report immediately using centralized Excel utilities"""
	try:
		from openpyxl.utils import get_column_letter
		
		# Prepare sheet data
		sheets_data = {
			"Cumulative OCR Report": {
				"headers": ocr_data.get("headers", []),
				"data": ocr_data.get("data", []),
				"special_formatting": {
					"highlighted_columns": {
						"columns": list(range(16, 23)),  # Refined columns (P-V)
						"color": "FFF2CC",
						"start_row": 2
					},
					"column_widths": {
						"E": 30  # Text Data column
						# Refined columns width will be set dynamically
					}
				}
			}
		}
		
		# Create workbook using centralized function
		workbook_result = create_styled_excel_workbook(sheets_data, "cumulative_ocr_report")
		
		if not workbook_result.get("success"):
			return {"success": False, "message": f"Error creating workbook: {workbook_result.get('error')}"}
		
		# Create download response
		download_result = create_download_response(
			workbook_result["workbook"], 
			workbook_result["filename"]
		)
		
		if download_result.get("success"):
			return {
				"success": True,
				"message": f"Report generated with {ocr_data.get('total_records', 0)} records including refined analysis",
				"download_triggered": True,
				"records_count": ocr_data.get("total_records", 0)
			}
		else:
			return {"success": False, "message": f"Error creating download: {download_result.get('error')}"}
		
	except Exception as e:
		frappe.log_error(f"Error generating report with centralized functions: {str(e)}")
		return {"success": False, "message": str(e)}


def generate_large_report_background_centralized():
	"""Background job for large report generation using centralized utilities"""
	try:
		# Get all OCR data using centralized function  
		ocr_data = get_consolidated_ocr_data(include_refined=True, format="dict")
		
		if not ocr_data or len(ocr_data) == 0:
			frappe.publish_realtime(
				event='report_error',
				message={'title': 'OCR Report Error', 'message': 'No OCR data found'},
				user=frappe.session.user
			)
			return {"success": False, "message": "No OCR data found"}
		
		total_records = len(ocr_data)
		
		# Use the existing background function with centralized data
		return generate_large_report_background_with_data(ocr_data)
		
	except Exception as e:
		error_msg = f"Error in background report generation: {str(e)}"
		frappe.log_error(error_msg)
		
		# Send error notification
		frappe.publish_realtime(
			event='report_error',
			message={'title': 'OCR Report Error', 'message': f'Error generating report: {str(e)}'},
			user=frappe.session.user
		)
		
		return {"success": False, "message": error_msg}


def generate_large_report_background_with_data(ocr_data):
	"""Generate large report using provided OCR data"""
	try:
		from kgk_customisations.utils.excel_utils import save_excel_as_frappe_file
		from openpyxl import Workbook
		from openpyxl.styles import PatternFill, Font
		from openpyxl.utils import get_column_letter
		
		# Create Excel workbook
		wb = Workbook()
		ws = wb.active
		ws.title = "Cumulative OCR Report"
		
		# Add headers
		headers = [
			"Upload Date", "Sequence", "Created On", "Batch Name", "Text Data",
			"Lot ID 1", "Lot ID 2", "Sub Lot ID", "Result", "Color", 
			"Blue UV", "Yellow UV", "Brown", "Type", "Fancy Yellow",
			"Refined Result", "Refined Color", "Refined Blue UV", "Refined Brown",
			"Refined Yellow UV", "Refined Type", "Refined Fancy Yellow"
		]
		ws.append(headers)
		
		# Style headers with proper highlighting
		header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
		yellow_header_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
		yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
		header_font = Font(bold=True)
		
		# Style header row
		for i, cell in enumerate(ws[1], 1):
			cell.font = header_font
			if i > 15:  # Refined columns (columns 16-22)
				cell.fill = yellow_header_fill
			else:
				cell.fill = header_fill
		
		processed_count = 0
		
		# Process each record
		for item in ocr_data:
			row_data = [
				item.get("upload_date", ""),
				item.get("sequence", ""),
				item.get("created_on", ""),
				item.get("batch_name", ""),
				(item.get("text_data", "") or "")[:1000],  # Truncate for memory
				item.get("lot_id_1", ""),
				item.get("lot_id_2", ""),
				item.get("sub_lot_id", ""),
				item.get("result", ""),
				item.get("color", ""),
				item.get("blue_uv", ""),
				item.get("yellow_uv", ""),
				item.get("brown", ""),
				item.get("type", ""),
				item.get("fancy_yellow", "")
			]
			
			# Add refined columns
			row_data.extend([
				item.get("refined_result", ""),
				item.get("refined_color", ""),
				item.get("refined_blue_uv", ""),
				item.get("refined_brown", ""),
				item.get("refined_yellow_uv", ""),
				item.get("refined_type", ""),
				item.get("refined_fancy_yellow", "")
			])
			
			ws.append(row_data)
			processed_count += 1
			
			# Progress update
			if processed_count % 1000 == 0:
				print(f"Processed {processed_count} records...")
		
		# Apply yellow highlighting to all refined column data cells
		refined_start_col = 16  # "Refined Result" column
		refined_end_col = 22    # "Refined Fancy Yellow" column
		
		print("Applying refined column highlighting...")
		for row_idx in range(2, ws.max_row + 1):
			for col_idx in range(refined_start_col, refined_end_col + 1):
				cell = ws.cell(row=row_idx, column=col_idx)
				cell.fill = yellow_fill
		
		print(f"Highlighting applied to {ws.max_row - 1} data rows")
		
		# Set column widths
		for col_idx in range(1, len(headers) + 1):
			column_letter = get_column_letter(col_idx)
			if col_idx == 5:  # Text Data column
				ws.column_dimensions[column_letter].width = 30
			elif col_idx >= refined_start_col:  # Refined columns
				ws.column_dimensions[column_letter].width = 12
			else:
				ws.column_dimensions[column_letter].width = 15
		
		# Save using centralized function
		filename = f"cumulative_ocr_report_{frappe.utils.today().replace('-', '_')}.xlsx"
		file_result = save_excel_as_frappe_file(wb, filename, is_private=True)
		
		if file_result.get("success"):
			# Send notification to user
			frappe.publish_realtime(
				event='report_ready',
				message={
					'title': 'OCR Report Ready',
					'message': f'Your cumulative OCR report with {processed_count} records is ready for download.',
					'file_url': file_result["file_url"],
					'records_count': processed_count
				},
				user=frappe.session.user
			)
			
			return {"success": True, "message": f"Background report completed with {processed_count} records"}
		else:
			raise Exception(f"File saving failed: {file_result.get('error')}")
		
	except Exception as e:
		error_msg = f"Error in background report generation with data: {str(e)}"
		print(error_msg)
		import traceback
		traceback.print_exc()
		
		# Send error notification
		frappe.publish_realtime(
			event='report_error',
			message={
				'title': 'OCR Report Error',
				'message': f'Error generating report: {str(e)}'
			},
			user=frappe.session.user
		)
		
		return {"success": False, "message": error_msg}


def generate_small_report_immediate(total_count):
	"""Generate report immediately for smaller datasets with refined columns"""
	try:
		from openpyxl import Workbook
		from openpyxl.styles import PatternFill, Font
		from openpyxl.utils import get_column_letter
		import io
		
		# Get all uploads
		uploads = frappe.db.sql("""
			SELECT name, creation 
			FROM `tabOCR Data Upload` 
			WHERE docstatus != 2
		""", as_dict=True)
		
		if not uploads:
			return {"success": False, "message": "No OCR data uploads found"}
		
		# Create Excel workbook
		wb = Workbook()
		ws = wb.active
		ws.title = "Cumulative OCR Report"
		
		# Define headers including refined columns
		headers = [
			"Upload Date", "Sequence", "Created On", "Batch Name", "Text Data",
			"Lot ID 1", "Lot ID 2", "Sub Lot ID", "Result", "Color", 
			"Blue UV", "Yellow UV", "Brown", "Type", "Fancy Yellow",
			"Refined Result", "Refined Color", "Refined Blue UV", "Refined Brown",
			"Refined Yellow UV", "Refined Type", "Refined Fancy Yellow"
		]
		ws.append(headers)
		
		# Process data
		processed_count = 0
		for upload in uploads:
			upload_date = frappe.utils.formatdate(upload.creation) if upload.creation else ""
			
			# Get items for this upload
			items = frappe.db.sql("""
				SELECT sequence, created_on, batch_name, text_data, lot_id_1, lot_id_2, 
					   sub_lot_id, result, color, blue_uv, yellow_uv, brown, type, fancy_yellow
				FROM `tabOCR Data Item` 
				WHERE parent = %s
			""", upload.name, as_dict=True)
			
			for item in items:
				# Base row data
				row = [
					upload_date,
					item.sequence or "",
					frappe.utils.formatdate(item.created_on) if item.created_on else "",
					item.batch_name or "",
					(item.text_data or "")[:500],  # Truncate for memory
					item.lot_id_1 or "",
					item.lot_id_2 or "",
					item.sub_lot_id or "",
					item.result or "",
					item.color or "",
					item.blue_uv or "",
					item.yellow_uv or "",
					item.brown or "",
					item.type or "",
					item.fancy_yellow or ""
				]
				
				# Add refined columns
				if item.text_data and item.text_data.strip() and len(item.text_data) < 5000:
					try:
						refined_data = extract_ocr_fields_from_text(item.text_data)
						row.extend([
							refined_data.get("Result", ""),
							refined_data.get("Color", ""),
							refined_data.get("Blue UV", ""),
							refined_data.get("Brown", ""),
							refined_data.get("Yellow UV", ""),
							refined_data.get("Type", ""),
							refined_data.get("Fancy Yellow", "")
						])
					except Exception as e:
						print(f"Error processing OCR for item: {str(e)}")
						# Add empty refined columns if extraction fails
						row.extend(["", "", "", "", "", "", ""])
				else:
					# Add empty refined columns if no OCR data
					row.extend(["", "", "", "", "", "", ""])
				
				ws.append(row)
				processed_count += 1
		
		# Style the headers
		header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
		header_font = Font(bold=True)
		
		for cell in ws[1]:
			cell.fill = header_fill
			cell.font = header_font
		
		# Highlight refined columns with yellow background
		yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
		
		# Refined columns start at column 16 (P) and go to column 22 (V)
		refined_start_col = 16  # "Refined Result" column
		refined_end_col = 22    # "Refined Fancy Yellow" column
		
		# Apply yellow highlighting to refined column headers
		for col_idx in range(refined_start_col, refined_end_col + 1):
			cell = ws.cell(row=1, column=col_idx)
			cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Darker yellow for headers
			cell.font = header_font
		
		# Apply yellow highlighting to all refined column data cells
		for row_idx in range(2, ws.max_row + 1):
			for col_idx in range(refined_start_col, refined_end_col + 1):
				cell = ws.cell(row=row_idx, column=col_idx)
				cell.fill = yellow_fill
		
		# Set column widths
		for col_idx in range(1, len(headers) + 1):
			column_letter = get_column_letter(col_idx)
			if col_idx == 5:  # Text Data column
				ws.column_dimensions[column_letter].width = 30
			elif col_idx >= refined_start_col:  # Refined columns
				ws.column_dimensions[column_letter].width = 12
			else:
				ws.column_dimensions[column_letter].width = 15
		
		# Save and return
		excel_buffer = io.BytesIO()
		wb.save(excel_buffer)
		excel_buffer.seek(0)
		
		filename = f"cumulative_ocr_report_{frappe.utils.today().replace('-', '_')}.xlsx"
		
		# Direct download
		frappe.local.response.filename = filename
		frappe.local.response.filecontent = excel_buffer.getvalue()
		frappe.local.response.type = "download"
		frappe.local.response.headers = {
			"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			"Content-Disposition": f"attachment; filename={filename}"
		}
		
		return {
			"success": True,
			"message": f"Report generated with {processed_count} records including refined analysis",
			"download_triggered": True,
			"records_count": processed_count
		}
		
	except Exception as e:
		error_msg = f"Error generating immediate report: {str(e)}"
		print(error_msg)
		import traceback
		traceback.print_exc()
		return {"success": False, "message": error_msg}


def generate_large_report_background():
	"""Background job for large report generation"""
	try:
		import pandas as pd
		from openpyxl import Workbook
		from openpyxl.styles import PatternFill, Font
		import io
		
		# Get all uploads
		uploads = frappe.db.sql("""
			SELECT name, creation 
			FROM `tabOCR Data Upload` 
			WHERE docstatus != 2
		""", as_dict=True)
		
		# Create Excel workbook
		wb = Workbook()
		ws = wb.active
		ws.title = "Cumulative OCR Report"
		
		# Add headers
		headers = [
			"Upload Date", "Sequence", "Created On", "Batch Name", "Text Data",
			"Lot ID 1", "Lot ID 2", "Sub Lot ID", "Result", "Color", 
			"Blue UV", "Yellow UV", "Brown", "Type", "Fancy Yellow",
			"Refined Result", "Refined Color", "Refined Blue UV", "Refined Brown",
			"Refined Yellow UV", "Refined Type", "Refined Fancy Yellow"
		]
		ws.append(headers)
		
		# Style headers with proper highlighting
		header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
		yellow_header_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Darker yellow for headers
		yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
		header_font = Font(bold=True)
		
		# Style header row
		for i, cell in enumerate(ws[1], 1):
			cell.font = header_font
			if i > 15:  # Refined columns (columns 16-22)
				cell.fill = yellow_header_fill
			else:
				cell.fill = header_fill
		
		processed_count = 0
		
		# Process each upload
		for upload in uploads:
			upload_date = frappe.utils.formatdate(upload.creation) if upload.creation else ""
			
			# Process in batches
			offset = 0
			batch_size = 500
			
			while True:
				items = frappe.db.sql("""
					SELECT sequence, created_on, batch_name, text_data, lot_id_1, lot_id_2, 
						   sub_lot_id, result, color, blue_uv, yellow_uv, brown, type, fancy_yellow
					FROM `tabOCR Data Item` 
					WHERE parent = %s
					LIMIT %s OFFSET %s
				""", (upload.name, batch_size, offset), as_dict=True)
				
				if not items:
					break
				
				for item in items:
					row_data = [
						upload_date,
						item.sequence or "",
						frappe.utils.formatdate(item.created_on) if item.created_on else "",
						item.batch_name or "",
						(item.text_data or "")[:1000],
						item.lot_id_1 or "",
						item.lot_id_2 or "",
						item.sub_lot_id or "",
						item.result or "",
						item.color or "",
						item.blue_uv or "",
						item.yellow_uv or "",
						item.brown or "",
						item.type or "",
						item.fancy_yellow or ""
					]
					
					# Add refined columns
					if item.text_data and len(item.text_data) < 3000:
						try:
							refined_data = extract_ocr_fields_from_text(item.text_data)
							row_data.extend([
								refined_data.get("Result", ""),
								refined_data.get("Color", ""),
								refined_data.get("Blue UV", ""),
								refined_data.get("Brown", ""),
								refined_data.get("Yellow UV", ""),
								refined_data.get("Type", ""),
								refined_data.get("Fancy Yellow", "")
							])
						except:
							row_data.extend(["", "", "", "", "", "", ""])
					else:
						row_data.extend(["", "", "", "", "", "", ""])
					
					ws.append(row_data)
					processed_count += 1
				
				offset += batch_size
				
				# Progress update
				if processed_count % 1000 == 0:
					print(f"Processed {processed_count} records...")
		
		# Apply yellow highlighting to all refined column data cells
		yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
		refined_start_col = 16  # "Refined Result" column
		refined_end_col = 22    # "Refined Fancy Yellow" column
		
		print("Applying refined column highlighting...")
		for row_idx in range(2, ws.max_row + 1):
			for col_idx in range(refined_start_col, refined_end_col + 1):
				cell = ws.cell(row=row_idx, column=col_idx)
				cell.fill = yellow_fill
		
		print(f"Highlighting applied to {ws.max_row - 1} data rows")
		
		# Save file
		excel_buffer = io.BytesIO()
		wb.save(excel_buffer)
		excel_buffer.seek(0)
		
		filename = f"cumulative_ocr_report_{frappe.utils.today().replace('-', '_')}.xlsx"
		
		# Save as file document
		file_doc = frappe.get_doc({
			"doctype": "File",
			"file_name": filename,
			"is_private": 1,
			"folder": "Home",
			"content": excel_buffer.getvalue()
		})
		file_doc.insert()
		frappe.db.commit()
		
		# Send notification to user
		frappe.publish_realtime(
			event='report_ready',
			message={
				'title': 'OCR Report Ready',
				'message': f'Your cumulative OCR report with {processed_count} records is ready for download.',
				'file_url': file_doc.file_url,
				'records_count': processed_count
			},
			user=frappe.session.user
		)
		
		return {"success": True, "message": f"Background report completed with {processed_count} records"}
		
	except Exception as e:
		error_msg = f"Error in background report generation: {str(e)}"
		print(error_msg)
		import traceback
		traceback.print_exc()
		
		# Send error notification
		frappe.publish_realtime(
			event='report_error',
			message={
				'title': 'OCR Report Error',
				'message': f'Error generating report: {str(e)}'
			},
			user=frappe.session.user
		)
		
		return {"success": False, "message": error_msg}
